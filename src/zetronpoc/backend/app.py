#!/usr/bin/env python3
"""
app.py - ZetronPOC v2.0 - API REST + servidor de estaticos.
http.server puro (sin Flask) para maxima portabilidad. Puerto 8080.

Todas las mutaciones (login, envio, CRUD, apply, pbx, cola, db, config)
registran auditoria (tabla auditoria) y log centralizado (tabla logs).
El repo es la unica fuente de verdad.
"""
import os, sys, json, subprocess, io, time, csv, re, urllib.parse
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

APP_DIR = os.environ.get("ZETRONPOC_DIR", "/opt/zetronpoc")
sys.path.insert(0, APP_DIR)
sys.path.insert(0, os.path.join(APP_DIR, "database"))
import db_manager as db
import vpn as vpnmod

HOST, PORT = "0.0.0.0", 8080
FRONT = os.path.join(APP_DIR, "frontend")
ALLOWED_PBX = re.compile(r'^(pjsip show|pjsip send|pjsip unregister|pjsip reload|core show|core restart|'
                         r'dialplan|module show|module reload|sip show|sip reload|iax2 show|reload)\b', re.I)

def jok(handler, data, code=200):
    body = json.dumps(data).encode("utf-8")
    handler.send_response(code)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)

def jtext(handler, text, code=200, ct="text/plain; charset=utf-8"):
    body = text.encode("utf-8")
    handler.send_response(code)
    handler.send_header("Content-Type", ct)
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)

def serve_file(handler, path, ct):
    if not os.path.exists(path):
        jtext(handler, "no encontrado", 404); return
    with open(path, "rb") as f: data = f.read()
    handler.send_response(200)
    handler.send_header("Content-Type", ct)
    handler.send_header("Content-Length", str(len(data)))
    handler.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
    handler.end_headers()
    handler.wfile.write(data)

def need_auth(handler):
    tok = handler.headers.get("Authorization", "").replace("Bearer ", "")
    return db.verificar_token(tok)

def _tok(handler):
    return handler.headers.get("Authorization", "").replace("Bearer ", "")

def _ip(handler):
    try: return handler.client_address[0] or ""
    except Exception:
        return ""

def aud(handler, accion, entidad, eid="", detalle=""):
    """Audita una mutacion. Nunca rompe la peticion."""
    try:
        db.registrar_auditoria(db.token_user(_tok(handler)), accion, entidad,
                               str(eid or ""), str(detalle or "")[:240], _ip(handler))
    except Exception:
        pass

def evlog(handler, nivel, origen, mensaje):
    """Log centralizado en la tabla logs."""
    try:
        db.registrar_log(nivel, origen, mensaje)
    except Exception:
        pass

def parse_rows_from_upload(filename, raw):
    rows = []
    if filename.lower().endswith(".csv"):
        text = raw.decode("utf-8", errors="replace")
        sniffer = csv.Sniffer()
        try: delim = sniffer.sniff(text).delimiter
        except Exception: delim = ","
        for r in csv.DictReader(io.StringIO(text), delimiter=delim):
            rows.append({k.strip(): v for k, v in r.items() if k})
    else:
        try:
            import openpyxl
        except ImportError:
            return None, "openpyxl no instalado"
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        keys = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
        for row in ws.iter_rows(min_row=2):
            vals = [str(c.value).strip() if c.value is not None else "" for c in row]
            rows.append(dict(zip(keys, vals)))
    return rows, None

def build_xlsx(headers, rows, sheet="Sheet1"):
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = sheet
    ws.append(headers)
    for r in rows: ws.append(r)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def serve_xlsx(handler, filename, data):
    handler.send_response(200)
    handler.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    handler.send_header("Content-Disposition", 'attachment; filename="%s"' % filename)
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)

def pbx_run(cmd):
    if not ALLOWED_PBX.match(cmd):
        return {"error": "comando no permitido"}
    try:
        out = subprocess.run(["asterisk", "-rx", cmd], capture_output=True, text=True, timeout=15)
        return {"salida": (out.stdout or out.stderr or "")[:8000]}
    except Exception as e:
        return {"error": str(e)}

def ext_status():
    out = {}
    try:
        r = subprocess.run(["asterisk", "-rx", "pjsip show registrations"], capture_output=True, text=True, timeout=10)
        for line in (r.stdout or "").splitlines():
            m = re.match(r'\s*([\w\-]+)/.*?\s+(\w+)\s', line)
            if m and m.group(2) in ("Registered", "Rejected", "Unregistered", "Trying", "Auth", "Sent"):
                key = m.group(1)
                if key.startswith("reg-"):
                    key = key[4:]
                out[key] = m.group(2)
    except Exception:
        pass
    return out

def diagnose():
    ip = db.get_config("hospital_pbx_ip", "")
    puerto = db.get_config("hospital_pbx_port", "5060")
    pasos = []
    pr = subprocess.run(["ping", "-c", "2", "-W", "2", ip], capture_output=True, text=True, timeout=10) if ip else None
    pasos.append({"paso": "Ping a central %s" % ip, "ok": bool(pr and pr.returncode == 0),
                  "salida": (pr.stdout or pr.stderr or "sin IP")[-400:]})
    sr = subprocess.run(["bash", "-c", "timeout 3 bash -c 'echo > /dev/tcp/%s/%s' 2>&1" % (ip, puerto)],
                        capture_output=True, text=True) if ip else None
    pasos.append({"paso": "Puerto SIP %s:%s" % (ip, puerto), "ok": bool(sr and sr.returncode == 0),
                  "salida": "OK" if sr and sr.returncode == 0 else "cerrado/inaccesible"})
    conf = ""
    try:
        with open("/etc/asterisk/pjsip.conf") as f: conf = f.read()[:4000]
    except Exception as e:
        conf = "No existe: %s" % e
    reg = pbx_run("pjsip show registrations")["salida"]
    return {"ip": ip, "puerto": puerto, "pasos": pasos,
            "pjsip_conf": conf, "registros": reg,
            "transportes": pbx_run("pjsip show transports")["salida"],
            "log_pjsip": ""}

# ===================== DIAGNOSTICO POCSAG EN VIVO =====================
MMDVM_LOG_DIR = "/var/log/mmdvm"

def diag_mqtt_response():
    """Captura on-demand las respuestas OK/KO que publica MMDVMHost en
    <name>/response. Lanza mosquitto_sub con timeout 2s y -C 5 (max 5 msgs)."""
    host = (db.get_config("mmdvm_mqtt_host", "127.0.0.1") or "127.0.0.1").strip() or "127.0.0.1"
    port = str(db.get_config("mmdvm_mqtt_port", "1883") or "1883").strip() or "1883"
    name = (db.get_config("mmdvm_mqtt_name", "host") or "host").strip() or "host"
    topic = "%s/response" % name
    try:
        r = subprocess.run(["timeout", "2", "mosquitto_sub", "-h", host, "-p", port,
                            "-t", topic, "-C", "5"], capture_output=True, text=True, timeout=4)
        lineas = [l for l in (r.stdout or "").splitlines() if l.strip()]
        return {"topic": topic, "host": host, "port": port, "lineas": lineas, "ok": True}
    except FileNotFoundError:
        return {"topic": topic, "host": host, "port": port, "lineas": [], "ok": False,
                "error": "mosquitto_sub no instalado"}
    except Exception as e:
        return {"topic": topic, "host": host, "port": port, "lineas": [], "ok": False,
                "error": str(e)[:200]}

def diag_mmdvm_log(lines=100):
    """Tail del MMDVM-*.log mas reciente filtrado a POCSAG/remote/NAK/Transmitted."""
    try:
        import glob
        files = sorted(glob.glob(os.path.join(MMDVM_LOG_DIR, "MMDVM-*.log")), reverse=True)
    except Exception:
        files = []
    if not files:
        # fallback: journalctl del servicio mmdvmhost
        try:
            r = subprocess.run(["journalctl", "-u", "mmdvmhost", "-n", str(int(lines)), "--no-pager"],
                               capture_output=True, text=True, timeout=5)
            lineas = [l for l in (r.stdout or "").splitlines()
                      if l.strip() and any(k in l.lower() for k in ("pocsag", "remote command", "nak", "transmitted", "page", "mqtt"))]
            return {"path": "journalctl -u mmdvmhost", "lineas": lineas, "ok": True,
                    "note": "no hay MMDVM-*.log en %s; usando journalctl" % MMDVM_LOG_DIR}
        except Exception as e:
            return {"path": MMDVM_LOG_DIR, "lineas": [], "ok": False,
                    "error": "no hay logs MMDVM-*.log en %s y journalctl fallo: %s" % (MMDVM_LOG_DIR, str(e)[:120])}
    path = files[0]
    try:
        r = subprocess.run(["bash", "-c", "tail -n %d %s | grep -Ei 'pocsag|remote command|nak|transmitted|page|mqtt'" % (int(lines), path)],
                           capture_output=True, text=True, timeout=3)
        lineas = [l for l in (r.stdout or "").splitlines() if l.strip()]
        return {"path": path, "lineas": lineas, "ok": True}
    except Exception as e:
        return {"path": path, "lineas": [], "ok": False, "error": str(e)[:200]}

def diag_dispatch_log(lines=50):
    """Tail del dispatch_mqtt.log para ver el payload exacto enviado."""
    path = os.path.join(APP_DIR, "logs", "dispatch_mqtt.log")
    if not os.path.exists(path):
        return {"path": path, "lineas": [], "ok": False, "error": "no existe el log (todavia no se despacho nada)"}
    try:
        with open(path, "r", errors="replace") as f:
            all_lines = f.readlines()
        lineas = [l.rstrip() for l in all_lines[-int(lines):] if l.strip()]
        return {"path": path, "lineas": lineas, "ok": True}
    except Exception as e:
        return {"path": path, "lineas": [], "ok": False, "error": str(e)[:200]}

def diag_config_check():
    """Parsea el MMDVM.ini real y valida valores criticos para Jumbospot."""
    import configparser
    ini_path = db.MMDVM_INI
    out = {"path": ini_path, "exists": os.path.exists(ini_path), "checks": []}
    if not out["exists"]:
        out["error"] = "MMDVM.ini no existe en %s (aplica config desde Parametros primero)" % ini_path
        return out
    cp = configparser.ConfigParser()
    try:
        cp.read(ini_path)
    except Exception as e:
        out["error"] = "no se pudo parsear: %s" % str(e)[:120]
        return out
    def g(section, key, default=""):
        try:
            return cp.get(section, key).strip()
        except Exception:
            return default
    dapnet = g("POCSAG Network", "Enable", "0")
    out["checks"].append({"k": "POCSAG Network (DAPNET) Enable", "v": dapnet or "0", "ok": (dapnet == "0"),
                          "hint": "debe ser 0. MMDVMHost usa [POCSAG Network] (con espacio), no [DAPNET]; si esta en 1, DAPNET transmite pages ajenos y mezcla basura"})
    pocsag_enable = g("POCSAG", "Enable", "0")
    out["checks"].append({"k": "POCSAG Enable", "v": pocsag_enable or "0", "ok": (pocsag_enable == "1"),
                          "hint": "debe ser 1; el baud POCSAG lo define el firmware del MMDVM, no el .ini"})
    # test_mode=1 es el #1 motivo de 'el pager no recibe nada': pocsag_handler.py
    # NO transmite por MMDVM (solo loguea 'modo test'). Es config de la BD, no del .ini.
    test_mode = db.get_config("test_mode", "0")
    out["checks"].append({"k": "Modo prueba (test_mode)", "v": test_mode or "0", "ok": (test_mode == "0"),
                          "hint": "CRITICO: si es 1, pocsag_handler.py NO transmite (solo loguea 'modo test') -> el pager NUNCA recibe. Pasar a 0 en Parametros > IVR > Modo prueba."})
    txinvert = g("Modem", "TXInvert", "0")
    out["checks"].append({"k": "Modem TXInvert", "v": txinvert, "ok": (txinvert == "1"),
                          "hint": "Jumbospot requiere 1 (polaridad FSK)"})
    pttinvert = g("Modem", "PTTInvert", "0")
    out["checks"].append({"k": "Modem PTTInvert", "v": pttinvert, "ok": (pttinvert == "1"),
                          "hint": "Jumbospot requiere 1"})
    txlevel = g("Modem", "TXLevel", "50")
    out["checks"].append({"k": "Modem TXLevel", "v": txlevel, "ok": True,
                          "hint": "si hay over-deviation, probar 25-35"})
    # version MMDVMHost — parsea la fecha (ej: "MMDVM-Host version 20260713 git #...")
    ver = "no disponible"
    for cand in ["/usr/local/bin/MMDVM-Host", "MMDVMHost", "MMDVM-Host"]:
        try:
            r = subprocess.run([cand, "-v"], capture_output=True, text=True, timeout=3)
            raw = ((r.stdout or "") + (r.stderr or "")).strip()
            if raw:
                ver = raw.splitlines()[0]
                break
        except Exception:
            continue
    _vmatch = re.search(r'(\d{8})', ver)
    _vok = True
    _vhint = "version: %s" % ver
    if _vmatch:
        _vdate = _vmatch.group(1)
        try:
            _vy = int(_vdate[:4])
            if _vy >= 2024:
                _vok = True
                _vhint = "version reciente (%s): soporta page y page_bcd por MQTT" % _vdate
            else:
                _vok = False
                _vhint = "version vieja (%s): puede no procesar page_bcd; actualizar con instalador_mmdvm.sh" % _vdate
        except Exception:
            pass
    out["checks"].append({"k": "MMDVMHost version", "v": ver, "ok": _vok, "hint": _vhint})
    # [MQTT] section — si no esta habilitado o el Name no coincide, dispatch_mqtt publica al vacio
    mqtt_en = g("MQTT", "Enable", "0")
    out["checks"].append({"k": "MQTT Enable", "v": mqtt_en or "0", "ok": (mqtt_en == "1"),
                          "hint": "debe ser 1: si esta en 0, MMDVMHost ignora los page de dispatch_mqtt (no hay respuesta en host/response)"})
    mqtt_name = g("MQTT", "Name", "")
    out["checks"].append({"k": "MQTT Name", "v": mqtt_name or "(ausente)", "ok": (mqtt_name == "host"),
                          "hint": "debe ser 'host': dispatch_mqtt publica a host/command; si difiere, el comando no llega"})
    mqtt_host = g("MQTT", "Host", "")
    out["checks"].append({"k": "MQTT Host", "v": mqtt_host or "(ausente)", "ok": (mqtt_host == "127.0.0.1"),
                          "hint": "debe apuntar al broker local (127.0.0.1)"})
    mqtt_port = g("MQTT", "Port", "")
    out["checks"].append({"k": "MQTT Port", "v": mqtt_port or "(ausente)", "ok": (mqtt_port == "1883"),
                          "hint": "puerto del broker mosquitto (1883)"})
    # estado del broker mosquitto: si no esta activo, dispatch_mqtt publica al vacio
    broker_status = "(no disponible)"
    try:
        rb = subprocess.run(["systemctl", "is-active", "mosquitto"], capture_output=True, text=True, timeout=3)
        broker_status = (rb.stdout or "").strip() or (rb.stderr or "").strip() or "?"
    except Exception as e:
        broker_status = "error: %s" % str(e)[:80]
    out["checks"].append({"k": "Mosquitto broker", "v": broker_status, "ok": (broker_status == "active"),
                          "hint": "debe estar 'active'; si no, sudo systemctl enable --now mosquitto"})
    # [RemoteControl] Enable=1 es OBLIGATORIO: es lo que hace que MMDVMHost se
    # suscriba al topic MQTT <Name>/command (host/command). Sin esto, dispatch_mqtt
    # publica al vacio: no aparece "remote command" en el log ni OK/KO en host/response.
    # (El socket TCP 7642 es solo un proxy de que RemoteControl esta activo; dispatch
    # NO usa ese socket, publica por MQTT.)
    rc_enable = g("Remote Control", "Enable", "0")
    out["checks"].append({"k": "Remote Control Enable", "v": rc_enable or "0", "ok": (rc_enable == "1"),
                          "hint": "debe ser 1 (seccion [Remote Control] CON espacio; MMDVMHost no reconoce [RemoteControl]): sin esto NO se suscribe a host/command y los page se pierden. Fix: re-Aplicar config MMDVM o re-ejecutar instalador_mmdvm.sh"})
    rc_port = g("Remote Control", "Port", "7642") or "7642"
    rc_state = "(no disponible)"
    try:
        import socket as _sock
        p = int(rc_port)
        s = _sock.create_connection(("127.0.0.1", p), timeout=2)
        s.close()
        rc_state = "escuchando en 127.0.0.1:%s" % p
    except Exception as e:
        rc_state = "no conecta: %s" % str(e)[:80]
    out["checks"].append({"k": "RemoteControl socket", "v": rc_state, "ok": rc_state.startswith("escuchando"),
                          "hint": "proxy de que [RemoteControl] esta activo; si no conecta, el .ini activo tiene Enable=0 o falta la seccion -> re-Aplicar config MMDVM"})
    # Check DEFINITIVO: cuantas suscripciones MQTT registro MMDVMHost al arranque.
    # MMDVM-Host.cpp suscribe "display-in" siempre, y "command" SOLO si
    # getRemoteControlEnabled()==true. on_subscribe imprime una linea por topic
    # concedido -> 2 lineas = RemoteControl activo; 1 linea = solo display-in
    # -> el .ini que corre MMDVMHost NO tiene [RemoteControl] Enable=1 (service
    # stale apuntando a otro .ini). Mas confiable que el socket TCP 7642 (algunos
    # builds no abren ese puerto).
    sub_count = 0
    try:
        import glob as _glob
        _files = sorted(_glob.glob(os.path.join(MMDVM_LOG_DIR, "MMDVM-*.log")), reverse=True)
        if _files:
            with open(_files[0], "r", errors="replace") as _f:
                _lines = _f.readlines()[-80:]
            sub_count = sum(1 for _l in _lines if "on_subscribe:" in _l)
    except Exception:
        pass
    out["checks"].append({"k": "Suscripciones MQTT (on_subscribe)", "v": "%d (esperadas 2)" % sub_count,
                          "ok": sub_count >= 2,
                          "hint": "2=display-in+command (OK). 1=solo display-in -> RemoteControl OFF en el .ini vivo -> service stale. Fix: sobreescribir /etc/systemd/system/mmdvmhost.service con ExecStart apuntando a /opt/zetronpoc/mmdvm/MMDVM.ini + Aplicar a la placa."})
    return out

def diag_test_page(cap, mensaje, funcion="alphanumeric"):
    """Dispara un page real via dispatch_mqtt.py para observar la respuesta OK/KO.
    funcion='numeric' -> page_bcd (func 0, BCD) para pagers numericos (NP88).
    funcion='alphanumeric' -> page (func 3, packASCII) para pagers alfanumericos.
    Si MMDVMHost es viejo (sin PAGE_BCD), page_bcd responde 'Invalid remote command'."""
    script = os.path.join(APP_DIR, "agi", "dispatch_mqtt.py")
    if not os.path.exists(script):
        return {"ok": False, "error": "dispatch_mqtt.py no encontrado en %s" % script}
    try:
        cap_int = int(str(cap).split(",")[0])
    except (ValueError, TypeError):
        return {"ok": False, "error": "cap invalido: %s" % str(cap)[:80]}
    bcd = (str(funcion).strip().lower() == "numeric")
    try:
        env = dict(os.environ, ZETRONPOC_DIR=APP_DIR)
        argv = [sys.executable, script]
        if bcd:
            argv.append("--bcd")
        argv += [str(cap_int), str(mensaje or "TEST"), "1200"]
        r = subprocess.run(argv, capture_output=True, text=True, timeout=15, env=env)
        return {"ok": r.returncode == 0, "stdout": (r.stdout or "").strip()[:500],
                "stderr": (r.stderr or "").strip()[:500], "rc": r.returncode, "bcd": bcd,
                "comando": "page_bcd" if bcd else "page"}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}

class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def _body(self):
        # Cachea el body: _post llama _json() para todos los endpoints y luego
        # los de upload (import / db restore) llaman _body() de nuevo. Sin cache,
        # la 2da lectura devuelve b"" (stream ya consumido) -> import 0 filas.
        if not hasattr(self, "_cached_body"):
            n = int(self.headers.get("Content-Length", 0))
            self._cached_body = self.rfile.read(n) if n else b""
        return self._cached_body

    def _json(self):
        try: return json.loads(self._body() or "{}")
        except Exception: return {}

    def do_GET(self):
        try: return self._get()
        except Exception as e:
            import traceback; tb=traceback.format_exc()
            try: return jok(self, {"error": str(e), "traceback": tb[-2000:]}, 500)
            except Exception: pass
    def _get(self):
        u = urllib.parse.urlparse(self.path); p = u.path; q = urllib.parse.parse_qs(u.query)
        if p == "/" or p == "/index.html": return serve_file(self, os.path.join(FRONT, "public.html"), "text/html; charset=utf-8")
        if p == "/admin" or p == "/admin.html": return serve_file(self, os.path.join(FRONT, "admin.html"), "text/html; charset=utf-8")
        if p == "/api/health": return jok(self, {"status": "ok", "ts": int(time.time())})
        if p == "/api/diag":
            out = {"db_path": db.DEFAULT_DB, "exists": os.path.exists(db.DEFAULT_DB)}
            tables = ["config","extensiones","pagers","grupos","grupo_miembros","bitacora","cola_envios","plantillas","envios_programados","auditoria","logs"]
            errs = []; counts = {}
            for t in tables:
                try:
                    with db.get_conn() as conn:
                        counts[t] = conn.execute("SELECT COUNT(*) FROM %s" % t).fetchone()[0]
                except Exception as e:
                    errs.append("%s: %s" % (t, str(e)[:160])); counts[t] = None
            out["counts"] = counts; out["errors"] = errs; out["db_ok"] = not errs
            return jok(self, out)
        if p == "/api/version": return jok(self, {"version": db.get_config("version", "2.0")})
        if p == "/api/server/time":
            import datetime as _dt
            return jok(self, {"datetime": _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                              "iso": _dt.datetime.now().isoformat(),
                              "date": _dt.date.today().isoformat(),
                              "timezone": time.tzname[0] if time.tzname else ""})
        if p == "/api/theme": return jok(self, db.all_config())
        if p == "/api/pagers": return jok(self, db.buscar_pagers(q.get("q", [""])[0]))
        if p == "/api/grupos": return jok(self, db.buscar_grupos(q.get("q", [""])[0]))
        if p == "/api/historial/public":
            return jok(self, db.historial({}, 100, 0))
        if p == "/api/enviar/estado":
            qid = int(q.get("id", ["0"])[0] or 0)
            st = db.estado_cola_id(qid)
            return jok(self, st or {"estado": "no_encontrado"})
        if p == "/api/plantillas": return jok(self, db.listar_plantillas())
        if p == "/api/login": return jtext(self, "use POST", 405)
        if not need_auth(self): return jok(self, {"error": "no autorizado"}, 401)

        if p == "/api/pagers/export":
            rows = db.buscar_pagers("")
            data = build_xlsx(["codigo","cap_code","nombre","apellido","area","baudios","funcion","descripcion"],
                [[r.get("codigo"), r.get("cap_code"), r.get("nombre") or "", r.get("apellido") or "",
                  r.get("area") or "", r.get("baudios"), r.get("funcion") or "", r.get("descripcion") or ""] for r in rows])
            return serve_xlsx(self, "pagers.xlsx", data)
        if p == "/api/grupos/export":
            rows = db.buscar_grupos("")
            data = build_xlsx(["codigo","nombre","baudios","cap_codes"],
                [[r.get("codigo"), r.get("nombre") or "", r.get("baudios"), ",".join(r.get("miembros") or [])] for r in rows])
            return serve_xlsx(self, "grupos.xlsx", data)
        if p == "/api/pagers/example":
            data = build_xlsx(["codigo","cap_code","nombre","apellido","area","baudios","funcion","descripcion"],
                [["1001","1234567","Juan","Perez","Guardia",512,"alphanumeric","Pager guardia turno manana"],
                 ["1002","7654321","Maria","Lopez","UCI",1200,"numeric","Numeric UCI"],
                 ["1003","1111111","Carlos","Soto","Codigos",512,"tone","Solo tono"]])
            return serve_xlsx(self, "ejemplo_pagers.xlsx", data)
        if p == "/api/grupos/example":
            data = build_xlsx(["codigo","nombre","baudios","cap_codes"],
                [["G01","Guardia general",512,"1234567,7654321"],
                 ["G02","UCI",1200,"1111111,2222222"]])
            return serve_xlsx(self, "ejemplo_grupos.xlsx", data)

        if p == "/api/diagnostico/mqtt_response": return jok(self, diag_mqtt_response())
        if p == "/api/diagnostico/mmdvm_log": return jok(self, diag_mmdvm_log(int(q.get("lines", ["100"])[0])))
        if p == "/api/diagnostico/dispatch_log": return jok(self, diag_dispatch_log(int(q.get("lines", ["50"])[0])))
        if p == "/api/diagnostico/config_check": return jok(self, diag_config_check())

        if p == "/api/vpn/status": return jok(self, vpnmod.status())
        if p == "/api/vpn/clients": return jok(self, vpnmod.list_clients())
        _mvpn = re.match(r'/api/vpn/clients/([A-Za-z0-9_\-]+)$', p)
        if _mvpn: return jok(self, vpnmod.get_client(_mvpn.group(1)))
        _mvpn_log = re.match(r'/api/vpn/clients/([A-Za-z0-9_\-]+)/connect_log$', p)
        if _mvpn_log: return jok(self, vpnmod.connect_log(_mvpn_log.group(1), int(q.get("lines", ["80"])[0])))
        if p == "/api/vpn/server/status": return jok(self, vpnmod.server_status())
        if p == "/api/vpn/logs": return jok(self, vpnmod.logs(q.get("kind", ["server"])[0], int(q.get("lines", ["80"])[0])))

        if p == "/api/changelog":
            return jok(self, db.listar_changelog(q.get("fecha_desde",[""])[0], q.get("fecha_hasta",[""])[0], int(q.get("limit",["200"])[0]), int(q.get("offset",["0"])[0])))
        if p == "/api/pager_cambios":
            return jok(self, db.listar_pager_cambios(q.get("desde",[""])[0], q.get("hasta",[""])[0], int(q.get("limit",["100"])[0]), int(q.get("offset",["0"])[0])))
        if p == "/api/pager_cambios/export":
            res = db.listar_pager_cambios(q.get("desde",[""])[0], q.get("hasta",[""])[0], 100000, 0)
            import json as _json
            rows_out = []
            for r in res.get("rows",[]):
                try: cams = _json.loads(r.get("campos_json") or "[]")
                except Exception: cams = []
                if not cams:
                    rows_out.append([r.get("fecha_hora"), r.get("codigo"), r.get("usuario"), r.get("cantidad"), "", "", ""])
                for c in cams:
                    rows_out.append([r.get("fecha_hora"), r.get("codigo"), r.get("usuario"), r.get("cantidad"),
                                     c.get("label") or c.get("campo",""), c.get("antes",""), c.get("despues","")])
            data = build_xlsx(["Fecha/Hora","Código","Usuario","Cantidad","Campo","Antes","Después"], rows_out, "Cambios pagers")
            return serve_xlsx(self, "cambios_pagers.xlsx", data)
        if p == "/api/dbadmin/tables": return jok(self, db.db_admin_tables())
        if p == "/api/dbadmin/select":
            return jok(self, db.db_admin_select(q.get("table",[""])[0], int(q.get("limit",["100"])[0]), int(q.get("offset",["0"])[0])))
        if p == "/api/dbadmin/export":
            r = db.db_admin_select(q.get("table",[""])[0], 100000, 0)
            out = io.StringIO(); w = csv.writer(out); w.writerow(r.get("columns", []))
            for row in r.get("rows", []): w.writerow([row.get(c,"") for c in r.get("columns",[])])
            return jtext(self, out.getvalue(), 200, "text/csv; charset=utf-8")
        if p == "/api/config": return jok(self, db.all_config())
        if p == "/api/mmdvm/config": return jok(self, {k: v for k, v in db.all_config().items() if k.startswith("mmdvm_")})
        if p == "/api/extensions": return jok(self, db.listar_extensiones())
        if p == "/api/extensions/status": return jok(self, ext_status())
        if p == "/api/config/hospital_installer":
            script, n = db.generar_instalador_hospital()
            if q.get("view", ["0"])[0] == "1":
                return jok(self, {"script": script, "internos": n})
            data = script.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/x-shellscript; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="config_hospital.sh"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers(); self.wfile.write(data); return
        if p == "/api/programados": return jok(self, db.listar_programados())
        if p == "/api/auditoria":
            return jok(self, db.listar_auditoria(int(q.get("limit", ["200"])[0]), int(q.get("offset", ["0"])[0])))
        if p == "/api/logs":
            tipo = q.get("tipo", ["api"])[0]
            # logs en BD (tabla logs) cuando tipo=db
            if tipo == "db":
                with db.get_conn() as conn:
                    rows = [dict(r) for r in conn.execute(
                        "SELECT fecha_hora,nivel,origen,mensaje FROM logs ORDER BY id DESC LIMIT ?",
                        (int(q.get("limit", ["300"])[0]),))]
                return jok(self, {"lineas": ["%s [%s] %s: %s" % (r["fecha_hora"], r["nivel"], r["origen"], r["mensaje"]) for r in rows], "path": "sqlite:logs"})
            return jok(self, db.leer_logs(tipo, int(q.get("limit", ["300"])[0])))
        if p == "/api/stats": return jok(self, db.estadisticas())
        if p == "/api/cola": return jok(self, db.listar_cola(q.get("estado", [None])[0], int(q.get("limit", ["200"])[0])))
        if p == "/api/cola/estado": return jok(self, db.estado_cola())
        if p == "/api/historial":
            f = {k: q[k][0] for k in q if k not in ("limit", "offset")}
            return jok(self, db.historial(f, int(q.get("limit", ["50"])[0]), int(q.get("offset", ["0"])[0])))
        if p == "/api/historial/export":
            f = {k: q[k][0] for k in q}
            res = db.historial(f, 10000, 0)["rows"]
            out = io.StringIO()
            w = csv.writer(out); w.writerow(["fecha_hora","interno","codigo","cap_code","mensaje","baudios","estado","obs"])
            for r in res: w.writerow([r.get("fecha_hora"), r.get("interno_origen"), r.get("codigo"), r.get("cap_code"), r.get("mensaje"), r.get("baudios"), r.get("estado"), r.get("observaciones")])
            return jtext(self, out.getvalue(), 200, "text/csv; charset=utf-8")
        if p == "/api/db/backup":
            bf = db.backup_db()
            with open(bf, "rb") as f: data = f.read()
            self.send_response(200); self.send_header("Content-Type", "application/octet-stream")
            self.send_header("Content-Disposition", 'attachment; filename="%s"' % os.path.basename(bf))
            self.send_header("Content-Length", str(len(data))); self.end_headers(); self.wfile.write(data); return
        if p == "/api/pbx": return jok(self, pbx_run(q.get("cmd", [""])[0]))
        if p == "/api/pbx/diagnose": return jok(self, diagnose())
        if p == "/api/mmdvm/status":
            bin_ok = os.path.exists("/usr/local/bin/MMDVM-Host")
            svc = "unknown"
            try:
                r = subprocess.run(["systemctl", "is-active", "mmdvmhost"], capture_output=True, text=True, timeout=5)
                svc = (r.stdout or "").strip() or "unknown"
            except Exception:
                pass
            return jok(self, {"installed": bin_ok, "service": svc, "binary": "/usr/local/bin/MMDVM-Host" if bin_ok else None})
        return jtext(self, "no encontrado", 404)

    def do_POST(self):
        try: return self._post()
        except Exception as e:
            import traceback; tb=traceback.format_exc()
            try: return jok(self, {"error": str(e), "traceback": tb[-2000:]}, 500)
            except Exception: pass
    def _post(self):
        u = urllib.parse.urlparse(self.path); p = u.path
        if p == "/api/login":
            d = self._json()
            user = d.get("user", "")
            try:
                tok = db.login_validar(user, d.get("pass", ""))
            except Exception as e:
                evlog(self, "error", "api", "login_validar fallo: %s" % str(e)[:120])
                return jok(self, {"error": "base de datos no disponible"}, 401)
            ok = bool(tok)
            try:
                db.registrar_auditoria(user or "?", "login", "auth", "", "ok" if ok else "credenciales invalidas", _ip(self))
            except Exception:
                pass
            evlog(self, "info" if ok else "warn", "api", "login user=%s %s" % (user, "ok" if ok else "fail"))
            if tok: return jok(self, {"token": tok})
            return jok(self, {"error": "credenciales invalidas"}, 401)
        if p == "/api/enviar":
            d = self._json()
            codigo = d.get("codigo", ""); mensaje = d.get("mensaje", ""); origen = d.get("origen", "web")
            res = db.enviar_mensaje(codigo, mensaje, origen)
            # auditar como sistema (endpoint publico) con el origen declarado
            try:
                db.registrar_auditoria(origen or "web", "enviar", "mensaje", str(res.get("id", "")),
                                        "codigo=%s estado=%s" % (codigo, res.get("status", "")), _ip(self))
            except Exception:
                pass
            evlog(self, "info", "api", "enviar codigo=%s estado=%s" % (codigo, res.get("status", "")))
            return jok(self, res)
        if not need_auth(self): return jok(self, {"error": "no autorizado"}, 401)
        d = self._json()
        if p == "/api/changelog":
            nid = db.crear_changelog(d)
            aud(self, "crear", "changelog", nid, d.get("titulo", ""))
            return jok(self, {"id": nid})
        if p == "/api/dbadmin/sql":
            res = db.db_admin_run_sql(d.get("sql", ""))
            aud(self, "sql", "dbadmin", "", (d.get("sql","") or "")[:120])
            return jok(self, res)
        if p == "/api/diagnostico/test_page":
            cap = d.get("cap", "1234567")
            msg = d.get("mensaje", "TEST")
            funcion = d.get("funcion", "alphanumeric")
            res = diag_test_page(cap, msg, funcion)
            aud(self, "test_page", "diagnostico", str(cap), "func=%s rc=%s ok=%s" % (funcion, res.get("rc", "-"), res.get("ok")))
            evlog(self, "info", "diag", "test_page cap=%s func=%s ok=%s" % (cap, funcion, res.get("ok")))
            return jok(self, res)
        if p == "/api/vpn/clients":
            res = vpnmod.create_client(d)
            aud(self, "crear", "vpn_client", d.get("name", ""), d.get("protocol", ""))
            evlog(self, "info", "vpn", "create client %s %s" % (d.get("name", ""), d.get("protocol", "")))
            return jok(self, res)
        if p == "/api/vpn/clients/up":
            res = vpnmod.up(d.get("name", ""))
            aud(self, "conectar", "vpn_client", d.get("name", ""), "ok=%s" % res.get("ok"))
            return jok(self, res)
        if p == "/api/vpn/clients/down":
            res = vpnmod.down(d.get("name", ""))
            aud(self, "desconectar", "vpn_client", d.get("name", ""), "ok=%s" % res.get("ok"))
            return jok(self, res)
        if p == "/api/vpn/server/apply":
            res = vpnmod.server_apply(d)
            aud(self, "aplicar", "vpn_server", d.get("protocol", ""), "ok=%s" % res.get("ok"))
            evlog(self, "info" if res.get("ok") else "error", "vpn", "server apply %s ok=%s" % (d.get("protocol", ""), res.get("ok")))
            return jok(self, res)
        if p == "/api/vpn/server/stop":
            res = vpnmod.server_stop(d.get("protocol", ""))
            aud(self, "detener", "vpn_server", d.get("protocol", ""), "")
            return jok(self, res)
        if p == "/api/vpn/logs":
            return jok(self, vpnmod.logs(d.get("kind", "server"), int(d.get("lines", 80) or 80)))
        if p == "/api/mmdvm/apply":
            for k, v in d.items():
                if k.startswith("mmdvm_"):
                    db.set_config(k, "" if v is None else str(v))
            ok, msg = db.generar_mmdvm_ini()
            if not ok:
                aud(self, "aplicar", "mmdvm", "", "fallo: %s" % msg)
                evlog(self, "error", "mmdvm", "apply fallo: %s" % msg)
                return jok(self, {"ok": False, "error": "No se pudo generar MMDVM.ini: %s" % msg})
            try:
                r = subprocess.run(["systemctl", "restart", "mmdvmhost"], capture_output=True, text=True, timeout=20)
            except FileNotFoundError:
                aud(self, "aplicar", "mmdvm", "", "systemctl no disponible")
                return jok(self, {"ok": False, "error": "MMDVM.ini generado, pero systemctl no esta disponible en este host."})
            except subprocess.TimeoutExpired:
                aud(self, "aplicar", "mmdvm", "", "timeout restart")
                return jok(self, {"ok": False, "error": "MMDVM.ini generado, pero el reinicio de mmdvmhost demoro demasiado."})
            if r.returncode != 0:
                svc_err = (r.stderr or r.stdout or "").strip()
                aud(self, "aplicar", "mmdvm", "", "restart fallo: %s" % svc_err[:120])
                evlog(self, "error", "mmdvm", "restart fallo: %s" % svc_err[:200])
                return jok(self, {"ok": False, "error": "MMDVM.ini generado en %s, pero fallo reiniciar el servicio 'mmdvmhost': %s" % (db.MMDVM_INI, svc_err or "verifique que MMDVMHost este instalado")})
            aud(self, "aplicar", "mmdvm", "", "ok")
            evlog(self, "info", "mmdvm", "apply ok en %s" % db.MMDVM_INI)
            return jok(self, {"ok": True, "salida": "MMDVM.ini generado en %s y servicio mmdvmhost reiniciado." % db.MMDVM_INI, "ini": db.MMDVM_INI})
        if p == "/api/mmdvm/install":
            url = "https://raw.githubusercontent.com/gatoambroggio/mediguard-os-copy/main/src/zetronpoc/instalador_mmdvm.sh"
            aud(self, "instalar", "mmdvm", "", "inicio")
            evlog(self, "info", "mmdvm", "install inicio")
            try:
                r = subprocess.run(["bash", "-c", "curl -fsSL %s | bash" % url],
                                   capture_output=True, text=True, timeout=600)
                out = (r.stdout or "") + (r.stderr or "")
                aud(self, "instalar", "mmdvm", "", "rc=%s" % r.returncode)
                evlog(self, "info" if r.returncode == 0 else "error", "mmdvm", "install rc=%s" % r.returncode)
                return jok(self, {"ok": r.returncode == 0, "salida": out[-12000:], "code": r.returncode})
            except subprocess.TimeoutExpired:
                aud(self, "instalar", "mmdvm", "", "timeout")
                return jok(self, {"ok": False, "error": "La instalacion demoro mas de 10 minutos. Revisa: journalctl -u mmdvmhost"})
            except Exception as e:
                aud(self, "instalar", "mmdvm", "", "excepcion: %s" % str(e)[:120])
                return jok(self, {"ok": False, "error": str(e)})
        if p == "/api/extensions":
            nid = db.crear_extension(d)
            aud(self, "crear", "extension", nid, d.get("numero", ""))
            return jok(self, {"id": nid})
        if p == "/api/extensions/aplicar":
            ok, msg = db.generar_pjsip_conf()
            if ok:
                subprocess.run(["asterisk", "-rx", "pjsip reload"], capture_output=True, timeout=10)
                subprocess.run(["asterisk", "-rx", "pjsip send register"], capture_output=True, timeout=10)
            aud(self, "aplicar", "extensiones", "", "ok" if ok else "fallo: %s" % msg)
            evlog(self, "info" if ok else "error", "pbx", "extensions/aplicar %s" % msg[:80])
            return jok(self, {"ok": ok, "salida": msg})
        if p == "/api/pagers":
            nid = db.crear_pager(d)
            aud(self, "crear", "pager", nid, d.get("codigo", ""))
            return jok(self, {"id": nid})
        if p == "/api/pagers/import":
            fn = urllib.parse.unquote(self.headers.get("X-Filename", "import.csv"))
            rows, err = parse_rows_from_upload(fn, self._body())
            if err: return jok(self, {"error": err})
            res = db.importar_pagers(rows)
            aud(self, "importar", "pagers", "", "ok=%s err=%s" % (res.get("importados"), res.get("errores")))
            return jok(self, res)
        if p == "/api/grupos":
            nid = db.crear_grupo(d)
            aud(self, "crear", "grupo", nid, d.get("codigo", ""))
            return jok(self, {"id": nid})
        if p == "/api/grupos/import":
            fn = urllib.parse.unquote(self.headers.get("X-Filename", "import.csv"))
            rows, err = parse_rows_from_upload(fn, self._body())
            if err: return jok(self, {"error": err})
            res = db.importar_grupos(rows)
            aud(self, "importar", "grupos", "", "ok=%s err=%s" % (res.get("importados"), res.get("errores")))
            return jok(self, res)
        if p == "/api/plantillas":
            nid = db.crear_plantilla(d)
            aud(self, "crear", "plantilla", nid, d.get("nombre", ""))
            return jok(self, {"id": nid})
        if p == "/api/programados":
            nid = db.crear_programado(d)
            aud(self, "crear", "programado", nid, d.get("codigo", ""))
            return jok(self, {"id": nid})
        if p == "/api/cola/reintentar":
            cid = int(d.get("id", 0))
            db.reintentar_cola(cid)
            aud(self, "reintentar", "cola", cid, "")
            return jok(self, {"ok": True})
        if p == "/api/cola/limpiar":
            db.limpiar_cola()
            aud(self, "limpiar", "cola", "", "")
            return jok(self, {"ok": True})
        if p == "/api/pbx/reload":
            r1 = pbx_run("pjsip reload"); r2 = pbx_run("dialplan reload")
            aud(self, "reload", "pbx", "", "pjsip+dialplan")
            evlog(self, "info", "pbx", "reload")
            return jok(self, {"salida": r1.get("salida", "") + "\n" + r2.get("salida", "")})
        if p == "/api/pbx/restart":
            r = subprocess.run(["systemctl", "restart", "asterisk"], capture_output=True, text=True, timeout=20)
            aud(self, "restart", "pbx", "", "")
            evlog(self, "warn", "pbx", "restart")
            return jok(self, {"salida": r.stdout or r.stderr or "reiniciado"})
        if p == "/api/pbx/force-register":
            r = pbx_run("pjsip send register")
            aud(self, "force-register", "pbx", "", "")
            return jok(self, {"salida": r.get("salida", "ok")})
        if p == "/api/pbx/unregister":
            r = pbx_run("pjsip unregister")
            aud(self, "unregister", "pbx", "", "")
            return jok(self, {"salida": r.get("salida", "ok")})
        if p == "/api/pbx/run":
            cmd = d.get("cmd", "")
            res = pbx_run(cmd)
            aud(self, "run", "pbx", "", cmd[:120])
            evlog(self, "info", "pbx", "run: %s" % cmd[:120])
            return jok(self, res)
        if p == "/api/db/backup-email":
            bf = db.backup_db()
            r = db.enviar_email(db.get_config("backup_email"), "Backup ZetronPOC", "Backup adjunto.", bf)
            aud(self, "backup-email", "db", "", "ok" if "ok" in r else "fallo")
            return jok(self, r if "ok" in r else {"error": r.get("error", "fallo")})
        if p == "/api/db/restore":
            data = self._body()
            db.restore_db(data)
            aud(self, "restore", "db", "", "")
            evlog(self, "warn", "db", "restore")
            return jok(self, {"ok": True})
        if p == "/api/smtp/test":
            email = d.get("email", db.get_config("backup_email"))
            r = db.enviar_email(email, "ZetronPOC - test SMTP", "Prueba OK")
            aud(self, "test", "smtp", "", email or "")
            return jok(self, r if "ok" in r else {"error": r.get("error", "fallo")})
        return jtext(self, "no encontrado", 404)

    def do_PUT(self):
        try: return self._put()
        except Exception as e:
            import traceback; tb=traceback.format_exc()
            try: return jok(self, {"error": str(e), "traceback": tb[-2000:]}, 500)
            except Exception: pass
    def _put(self):
        if not need_auth(self): return jok(self, {"error": "no autorizado"}, 401)
        u = urllib.parse.urlparse(self.path); p = u.path; d = self._json()
        if p == "/api/config":
            keys = list(d.keys())
            for k, v in d.items(): db.set_config(k, "" if v is None else str(v))
            aud(self, "guardar", "config", "", "keys=%s" % ",".join(keys)[:200])
            evlog(self, "info", "config", "guardar %d keys" % len(keys))
            # Si se guardaron claves mmdvm_*, regenerar MMDVM.ini y reiniciar mmdvmhost
            mmdvm_aplicado = False; mmdvm_msg = ""; mmdvm_error = ""
            if any(k.startswith("mmdvm_") for k in keys):
                ok, msg = db.generar_mmdvm_ini()
                if ok:
                    try:
                        r = subprocess.run(["systemctl", "restart", "mmdvmhost"], capture_output=True, text=True, timeout=20)
                        if r.returncode == 0:
                            mmdvm_aplicado = True; mmdvm_msg = "MMDVM.ini regenerado y mmdvmhost reiniciado"
                        else:
                            mmdvm_error = "fallo reiniciar mmdvmhost: %s" % (r.stderr or r.stdout or "")[:160]
                    except Exception as e:
                        mmdvm_error = "systemctl no disponible: %s" % str(e)[:120]
                else:
                    mmdvm_error = "fallo generar MMDVM.ini: %s" % msg
            return jok(self, {"ok": True, "mmdvm_aplicado": mmdvm_aplicado, "mmdvm_msg": mmdvm_msg, "mmdvm_error": mmdvm_error})
        m = re.match(r'/api/extensions/(\d+)$', p)
        if m:
            db.actualizar_extension(int(m.group(1)), d)
            aud(self, "actualizar", "extension", m.group(1), d.get("numero", ""))
            return jok(self, {"ok": True})
        m = re.match(r'/api/pagers/(\d+)$', p)
        if m:
            if d.get("origen") == "cambios":
                res = db.actualizar_pager_con_cambio(int(m.group(1)), d, db.token_user(_tok(self)))
                aud(self, "editar", "pager", m.group(1), "cambios=%d" % res.get("cambios", 0))
                return jok(self, res)
            db.actualizar_pager(int(m.group(1)), d)
            aud(self, "actualizar", "pager", m.group(1), d.get("codigo", ""))
            return jok(self, {"ok": True})
        m = re.match(r'/api/pagers/(\d+)/estado$', p)
        if m:
            db.toggle_pager(int(m.group(1)), int(d.get("activo", 1)))
            aud(self, "toggle", "pager", m.group(1), "activo=%s" % d.get("activo", 1))
            return jok(self, {"ok": True})
        m = re.match(r'/api/grupos/(\d+)$', p)
        if m:
            db.actualizar_grupo(int(m.group(1)), d)
            aud(self, "actualizar", "grupo", m.group(1), d.get("codigo", ""))
            return jok(self, {"ok": True})
        m = re.match(r'/api/plantillas/(\d+)$', p)
        if m:
            db.actualizar_plantilla(int(m.group(1)), d)
            aud(self, "actualizar", "plantilla", m.group(1), d.get("nombre", ""))
            return jok(self, {"ok": True})
        m = re.match(r'/api/programados/(\d+)$', p)
        if m:
            db.actualizar_programado(int(m.group(1)), d)
            aud(self, "actualizar", "programado", m.group(1), d.get("codigo", ""))
            return jok(self, {"ok": True})
        m = re.match(r'/api/vpn/clients/([A-Za-z0-9_\-]+)$', p)
        if m:
            d["name"] = d.get("name") or m.group(1)
            res = vpnmod.create_client(d)
            aud(self, "actualizar", "vpn_client", m.group(1), d.get("protocol", ""))
            evlog(self, "info", "vpn", "update client %s" % m.group(1))
            return jok(self, res)
        return jtext(self, "no encontrado", 404)

    def do_DELETE(self):
        try: return self._delete()
        except Exception as e:
            import traceback; tb=traceback.format_exc()
            try: return jok(self, {"error": str(e), "traceback": tb[-2000:]}, 500)
            except Exception: pass
    def _delete(self):
        if not need_auth(self): return jok(self, {"error": "no autorizado"}, 401)
        p = urllib.parse.urlparse(self.path).path
        m = re.match(r'/api/changelog/(\d+)$', p)
        if m:
            db.borrar_changelog(int(m.group(1)))
            aud(self, "borrar", "changelog", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/extensions/(\d+)$', p)
        if m:
            db.borrar_extension(int(m.group(1)))
            aud(self, "borrar", "extension", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/pagers/(\d+)$', p)
        if m:
            db.borrar_pager(int(m.group(1)))
            aud(self, "borrar", "pager", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/grupos/(\d+)$', p)
        if m:
            db.borrar_grupo(int(m.group(1)))
            aud(self, "borrar", "grupo", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/plantillas/(\d+)$', p)
        if m:
            db.borrar_plantilla(int(m.group(1)))
            aud(self, "borrar", "plantilla", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/programados/(\d+)$', p)
        if m:
            db.borrar_programado(int(m.group(1)))
            aud(self, "borrar", "programado", m.group(1), "")
            return jok(self, {"ok": True})
        m = re.match(r'/api/vpn/clients/([A-Za-z0-9_\-]+)$', p)
        if m:
            res = vpnmod.delete_client(m.group(1))
            aud(self, "borrar", "vpn_client", m.group(1), "")
            return jok(self, res)
        return jtext(self, "no encontrado", 404)

class Server(ThreadingHTTPServer):
    daemon_threads = True

def _auto_init_db():
    """Garantiza que la base y sus tablas existen al arrancar. Idempotente
    (schema.sql usa CREATE TABLE IF NOT EXISTS). Nunca bloquea el arranque."""
    try:
        db.init_db()
        print("[init] Base de datos verificada/creada.", flush=True)
    except Exception as e:
        print("[init] WARN: no se pudo inicializar la base: %s" % e, flush=True)

if __name__ == "__main__":
    try:
        os.makedirs(os.path.join(APP_DIR, "logs"), exist_ok=True)
        _auto_init_db()
        print("ZetronPOC v2.0 API en http://%s:%d" % (HOST, PORT), flush=True)
        Server((HOST, PORT), Handler).serve_forever()
    except Exception:
        import traceback
        traceback.print_exc()
        raise